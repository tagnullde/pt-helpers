#!/usr/bin/env python3
# ----------------------------------------------------------------------------- #
#               haste.py — fast inital recon and documentation tool             #
#                         by x41 and the praktikant  (v4.5)                     #
# ----------------------------------------------------------------------------- #
# Stages:
#   1. masscan     — full port discovery (1-65535)
#   2. parse       — extract hosts/ports from masscan XML
#   3. confirm     — full TCP connect per port, filters proxied/dead ports only confirmed-open ports proceed to nmap
#   4. nmap        — -sV -sC -O per host on confirmed ports only produces per-host XMLs + merged nmap_combined.xml
#   5. xlsx        — nmap_results.xlsx from combined XML (openpyxl)
#   6. eyewitness  — screenshots of web services (optional, if installed)
#
# Usage:
#   sudo python3 scanner.py <TARGET>
#   sudo python3 scanner.py --iL <FILE>
#   sudo python3 scanner.py --from-masscan-xml <masscan.xml>
#   sudo python3 scanner.py --from-nmap-xml <nmap_combined.xml>
# ----------------------------------------------------------------------------- #

import argparse
import csv
import os
import re
import shutil
import signal
import subprocess
import sys
import threading
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from ipaddress import ip_address, ip_network
from pathlib import Path
from threading import Lock

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ----------------------------------------------------------------------------- #
# Spinner progress indicator
# ----------------------------------------------------------------------------- #
_SPINNER_FRAMES = ["-", "\\", "|", "/"]


class Spinner:
    """
    Background-thread spinner that runs independently of task completion.
    The spinner animates continuously; n/total updates as futures complete.

    Usage:
        sp = Spinner(total)
        sp.start()
        ...
        sp.update()   # call each time a task finishes
        ...
        sp.stop("  [+] all done")   # stops spinner, prints result line
    """

    def __init__(self, total: int, interval: float = 0.1):
        self.total    = total
        self.interval = interval
        self._n       = 0
        self._frame   = 0
        self._label   = ""
        self._stop    = threading.Event()
        self._lock    = threading.Lock()
        self._thread  = threading.Thread(target=self._run, daemon=True)

    def start(self) -> "Spinner":
        self._thread.start()
        return self

    def update(self) -> None:
        with self._lock:
            self._n += 1

    def set_label(self, label: str) -> None:
        """Update status label — used in indeterminate mode (total=0)."""
        with self._lock:
            self._label = label

    def stop(self, final_line: str = "") -> None:
        if self._stop.is_set():
            return  # already stopped — safe to call multiple times
        self._stop.set()
        self._thread.join()
        # erase spinner line, print final result
        print(f"\r{' ' * 70}\r", end="", flush=True)
        if final_line:
            print(final_line)

    def _run(self) -> None:
        while not self._stop.is_set():
            with self._lock:
                n     = self._n
                label = self._label
                frame = _SPINNER_FRAMES[self._frame % len(_SPINNER_FRAMES)]
                self._frame += 1
            if self.total:
                suffix = f"{n}/{self.total}"
            else:
                suffix = label  # indeterminate — show label from set_label()
            print(f"\r  [ {frame} ] {suffix} ", end="", flush=True)
            self._stop.wait(self.interval)


# ----------------------------------------------------------------------------- #
# Helpers
# ----------------------------------------------------------------------------- #
def sort_ips(ips: list[str]) -> list[str]:
    return sorted(ips, key=lambda x: ip_address(x))


def estimate_hosts(target: str) -> int:
    try:
        net = ip_network(target, strict=False)
        n = net.num_addresses
        return max(1, n - 2) if n > 2 else n
    except ValueError:
        return 1


def estimate_hosts_from_file(path: str) -> int:
    total = 0
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            total += estimate_hosts(line)
    return total


def format_time(seconds: int) -> str:
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


def validate_target(target: str) -> None:
    try:
        net = ip_network(target, strict=False)
        if net.num_addresses == 1:
            print(f"[!] Target {target} is a single host (/32).")
            print(f"    masscan on a single host may miss ports due to self-routing load.")
            print(f"    For single hosts, nmap is more reliable.")
            answer = input("    Continue anyway? (y/n): ").strip().lower()
            if answer not in ("y", "yes"):
                print("[!] Aborted.")
                sys.exit(1)
    except ValueError:
        print(f"[!] Invalid target: {target}")
        sys.exit(1)


# ----------------------------------------------------------------------------- #
# Dependency checks
# ----------------------------------------------------------------------------- #
def check_dependencies(skip_masscan: bool, require_nmap: bool = True) -> dict[str, bool]:
    required = []
    if require_nmap:
        required.append("nmap")
    if not skip_masscan:
        required.append("masscan")

    for tool in required:
        if not shutil.which(tool):
            print(f"[!] required tool not found: {tool}")
            sys.exit(1)

    optional = {"eyewitness": False}
    for tool in optional:
        found = shutil.which(tool) is not None
        optional[tool] = found
        if not found:
            answer = input(f"[?] {tool} not installed — continue without it? (y/n): ").strip().lower()
            if answer not in ("y", "yes"):
                print("[!] Aborted.")
                sys.exit(1)

    return optional


# ----------------------------------------------------------------------------- #
# Stage 1: masscan
# ----------------------------------------------------------------------------- #
def run_masscan(
    rate:    int,
    target:  str,
    il_file: str,
    exclude: str,
    outfile: Path,
) -> None:
    cmd = [
        "masscan",
        "-p", "1-65535",
        "--rate", str(rate),
        "--retries", "1",
        "--open-only",
        "-oX", str(outfile),
    ]

    if il_file:
        cmd += ["-iL", il_file]
    else:
        cmd.append(target)

    if exclude:
        if os.path.isfile(exclude):
            cmd += ["--excludefile", exclude]
        else:
            cmd += ["--exclude", exclude]

    # Run masscan with stderr piped so we can parse its progress lines.
    # masscan writes status to stderr in the form:
    #   rate:  49983; found: 234; remaining: 00:01:22; waiting...
    # Spinner runs in indeterminate mode (total=0); label updates with live stats.
    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        text=True,
    )

    found_count = {"n": 0}
    sp = Spinner(total=0).start()  # indeterminate — shows label instead of n/total

    def read_stderr():
        # Accumulate state from whatever masscan emits — different lines
        # carry different fields, so we build the label from running totals.
        # masscan 1.3.2 status line format:
        #   rate:  49.83-kpps, 2.95%
        # may also emit: found: N; remaining: HH:MM:SS
        state = {"pct": "", "rate": "", "found": "0", "rem": ""}
        for line in proc.stderr:
            line = line.strip()
            m_pct   = re.search(r"([\d.]+)%", line)
            m_rate  = re.search(r"rate:\s*([\d.]+-\w+)", line)
            m_found = re.search(r"found:\s*(\d+)", line)
            m_rem   = re.search(r"remaining:\s*(\S+)", line)
            if m_pct:   state["pct"]   = m_pct.group(1) + "%"
            if m_rate:  state["rate"]  = m_rate.group(1)
            if m_found:
                state["found"] = m_found.group(1)
                found_count["n"] = int(m_found.group(1))
            if m_rem:   state["rem"]   = m_rem.group(1).rstrip(";")

            parts = []
            if state["pct"]:   parts.append(state["pct"])
            if state["rem"]:   parts.append(f"ETA: {state['rem']}")
            if state["rate"]:  parts.append(f"rate: {state['rate']}")
            parts.append(f"found: {state['found']}")
            sp.set_label(f"{'  '.join(parts):<55}")

    t = threading.Thread(target=read_stderr, daemon=True)
    t.start()
    try:
        proc.wait()
    except KeyboardInterrupt:
        proc.terminate()
        proc.wait()
        sp.stop()
        raise
    t.join()  # stderr pipe closes on masscan exit; thread drains immediately
    sp.stop(f"  [+] masscan done — {found_count['n']} open ports found")

    # Exit code 1 can mean partial run (Ctrl+C) or real error.
    # Check whether the XML exists and is parseable instead of trusting the code.
    if proc.returncode != 0:
        if not outfile.exists() or outfile.stat().st_size == 0:
            print(f"[!] masscan failed (exit {proc.returncode}) and produced no output")
            sys.exit(1)
        try:
            ET.parse(outfile)
        except ET.ParseError:
            print(f"[!] masscan failed (exit {proc.returncode}) and output is not valid XML")
            sys.exit(1)
        # XML exists and is valid — partial run is usable, continue


# ----------------------------------------------------------------------------- #
# Stage 2: parse masscan XML
# ----------------------------------------------------------------------------- #
def parse_masscan_xml(path: Path) -> dict[str, set[int]]:
    try:
        tree = ET.parse(path)
    except ET.ParseError as e:
        print(f"[!] Failed to parse masscan XML: {e}")
        sys.exit(1)

    root = tree.getroot()
    portmap: dict[str, set[int]] = defaultdict(set)

    for host in root.findall("host"):
        addr_el = host.find("address[@addrtype='ipv4']")
        if addr_el is None:
            continue
        ip = addr_el.get("addr")
        if not ip:
            continue
        for port_el in host.findall(".//port[@protocol='tcp']"):
            state_el = port_el.find("state[@state='open']")
            if state_el is None:
                continue
            portid = port_el.get("portid")
            if portid and portid.isdigit():
                portmap[ip].add(int(portid))

    return portmap


def write_parse_outputs(portmap: dict[str, set[int]], outdir: Path) -> Path:
    """Write masscan parse outputs. Returns portmap_path for summary use."""
    sorted_hosts = sort_ips(list(portmap.keys()))
    all_ports    = sorted(set(p for ports in portmap.values() for p in ports))
    # Raw masscan parse output — one row per host with all discovered ports
    debug_dir    = outdir / "debug"
    debug_dir.mkdir(exist_ok=True)
    portmap_path = debug_dir / "masscan_ports_map.csv"
    with open(portmap_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["host", "ports"])
        for ip in sorted_hosts:
            writer.writerow([ip, ",".join(str(p) for p in sorted(portmap[ip]))])

    with open(debug_dir / "masscan_alive_hosts.txt", "w") as f:
        for ip in sorted_hosts:
            f.write(ip + "\n")

    with open(debug_dir / "masscan_union_ports.txt", "w") as f:
        for port in all_ports:
            f.write(str(port) + "\n")

    lines = [
        "=== masscan parse summary ===",
        "",
        f"Alive hosts (masscan)  : {len(sorted_hosts)}",
        f"Unique ports           : {len(all_ports)}",
        "",
        "--- Per-host breakdown ---",
    ]
    for ip in sorted_hosts:
        ports_str = ",".join(str(p) for p in sorted(portmap[ip]))
        lines.append(f"  {ip:<20}  ports: {ports_str}")

    with open(debug_dir / "masscan_summary.txt", "w") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  [+] masscan: {len(sorted_hosts)} hosts, {len(all_ports)} unique ports")
    return portmap_path


# ----------------------------------------------------------------------------- #
# Stage 3: parallel nmap TCP connect confirmation
# ----------------------------------------------------------------------------- #
def parse_nmap_open_ports(xml_output: str) -> set[int]:
    confirmed = set()
    try:
        root = ET.fromstring(xml_output)
        for port_el in root.findall(".//port"):
            state_el = port_el.find("state[@state='open']")
            if state_el is not None:
                portid = port_el.get("portid")
                if portid and portid.isdigit():
                    confirmed.add(int(portid))
    except ET.ParseError:
        pass
    return confirmed


# Registry for confirm-stage nmap processes (populated by confirm_hosts)
_confirm_procs:      list[subprocess.Popen] = []
_confirm_procs_lock: Lock = Lock()


def confirm_single_host(
    ip:      str,
    ports:   set[int],
    timeout: int,
) -> tuple[str, set[int]]:
    ports_arg = ",".join(str(p) for p in sorted(ports))
    cmd = [
        "nmap",
        "-sT", "-Pn", "-n",
        "--open",
        "--max-retries", "0",
        "--host-timeout", f"{timeout}s",
        "-p", ports_arg,
        "-oX", "-",
        ip,
    ]
    try:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True)
        with _confirm_procs_lock:
            _confirm_procs.append(proc)
        try:
            stdout, _ = proc.communicate(timeout=timeout + 5)
        except subprocess.TimeoutExpired:
            proc.kill()
            remaining, _ = proc.communicate()
            # Use whatever nmap wrote before being killed if it's valid XML;
            # otherwise discard — partial/truncated XML is not reliable.
            try:
                if remaining:
                    ET.fromstring(remaining)
                    stdout = remaining
                else:
                    stdout = ""
            except ET.ParseError:
                stdout = ""
        finally:
            with _confirm_procs_lock:
                try:
                    _confirm_procs.remove(proc)
                except ValueError:
                    pass
        return ip, parse_nmap_open_ports(stdout)
    except Exception:
        return ip, set()


def confirm_hosts(
    portmap:  dict[str, set[int]],
    workers:  int,
    timeout:  int,
    outdir:   Path,
    debug:    bool = False,
) -> tuple[dict[str, set[int]], Path]:
    sorted_hosts = sort_ips(list(portmap.keys()))
    total        = len(sorted_hosts)
    confirmed_map: dict[str, set[int]] = {}
    dropped:       list[str] = []

    sp = Spinner(total).start() if not debug else None

    def task(ip: str) -> tuple[str, set[int]]:
        return confirm_single_host(ip, portmap[ip], timeout)

    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(task, ip): ip for ip in sorted_hosts}
            for future in as_completed(futures):
                ip, confirmed_ports = future.result()
                if sp is not None:
                    sp.update()
                if confirmed_ports:
                    confirmed_map[ip] = confirmed_ports
                    if debug:
                        conf_str = ",".join(str(p) for p in sorted(confirmed_ports))
                        print(f"  {ip:<20} ✓  {conf_str}")
                else:
                    dropped.append(ip)
                    if debug:
                        print(f"  {ip:<20} ✗  dropped")
    except KeyboardInterrupt:
        # kill any in-flight confirm nmap processes
        with _confirm_procs_lock:
            for proc in _confirm_procs:
                try:
                    proc.terminate()
                except Exception:
                    pass
        if sp is not None:
            sp.stop()
        raise

    confirmed_sorted    = sort_ips(list(confirmed_map.keys()))
    all_confirmed_ports = sorted(set(p for ports in confirmed_map.values() for p in ports))

    # Debug outputs
    debug_dir = outdir / "debug"
    debug_dir.mkdir(exist_ok=True)

    # Write confirmed ports map to debug dir
    confirmed_ports_path = debug_dir / "confirmed_ports_map.csv"
    with open(confirmed_ports_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["host", "ports"])
        for ip in confirmed_sorted:
            writer.writerow([ip, ",".join(str(p) for p in sorted(confirmed_map[ip]))])

    with open(debug_dir / "confirmed_hosts.txt", "w") as f:
        for ip in confirmed_sorted:
            f.write(ip + "\n")

    lines = [
        "=== confirmation summary ===",
        "",
        f"Hosts in (masscan)     : {total}",
        f"Hosts confirmed alive  : {len(confirmed_sorted)}",
        f"Hosts dropped          : {len(dropped)}",
        f"Unique confirmed ports : {len(all_confirmed_ports)}",
        "",
        "--- Confirmed per-host ---",
    ]
    for ip in confirmed_sorted:
        ports_str = ",".join(str(p) for p in sorted(confirmed_map[ip]))
        lines.append(f"  {ip:<20}  {ports_str}")
    if dropped:
        lines += ["", "--- Dropped (SYN proxy / not alive) ---"]
        for ip in sort_ips(dropped):
            lines.append(f"  {ip}")
    with open(debug_dir / "confirmed_summary.txt", "w") as f:
        f.write("\n".join(lines) + "\n")

    if sp is not None:
        sp.stop(f"  [+] confirmed: {len(confirmed_sorted)} alive, {len(dropped)} dropped")
    else:
        print(f"  [+] confirmed: {len(confirmed_sorted)} alive, {len(dropped)} dropped")
    return confirmed_map, confirmed_ports_path


# ----------------------------------------------------------------------------- #
# Stage 4: nmap deep dive
# ----------------------------------------------------------------------------- #
def _merge_nmap_xmls(xml_paths: list[Path], out_path: Path) -> None:
    if not xml_paths:
        # Write a minimal valid nmaprun so downstream stages have a parseable file
        with open(out_path, "w") as f:
            f.write('<?xml version="1.0"?>\n<nmaprun scanner="nmap" args="scanner.py combined" start="" version="" xmloutputversion="1.04">\n</nmaprun>\n')
        return

    host_blocks: list[str] = []
    base_attribs: dict     = {}

    for path in xml_paths:
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            if not base_attribs:
                base_attribs = root.attrib.copy()
            for host_el in root.findall("host"):
                host_blocks.append(ET.tostring(host_el, encoding="unicode"))
        except ET.ParseError:
            continue

    with open(out_path, "w") as f:
        f.write('<?xml version="1.0"?>\n')
        f.write(
            f'<nmaprun scanner="nmap" args="scanner.py combined" '
            f'start="{base_attribs.get("start", "")}" '
            f'version="{base_attribs.get("version", "")}" '
            f'xmloutputversion="{base_attribs.get("xmloutputversion", "1.04")}">\n'
        )
        for block in host_blocks:
            f.write(block + "\n")
        f.write("</nmaprun>\n")


def run_nmap_deep_dive(
    confirmed_map: dict[str, set[int]],
    workers:       int,
    timeout:       int,
    outdir:        Path,
    debug:         bool = False,
) -> tuple[Path, list[str]]:
    sorted_hosts = sort_ips(list(confirmed_map.keys()))
    total        = len(sorted_hosts)

    nmap_dir = outdir / "nmap"
    nmap_dir.mkdir(exist_ok=True)

    sp = Spinner(total).start() if not debug else None

    # Registry of all live nmap child processes so we can kill them on Ctrl+C
    _procs:      list[subprocess.Popen] = []
    _procs_lock: Lock = Lock()
    print_lock:  Lock = Lock()
    _abort       = threading.Event()

    def scan_host(ip: str) -> tuple[str, Path | None]:
        if _abort.is_set():
            return ip, None

        ports     = confirmed_map[ip]
        ports_arg = ",".join(str(p) for p in sorted(ports))
        safe_ip   = ip.replace(".", "_")
        # -oA produces .xml + .nmap + .gnmap
        oA_base   = nmap_dir / safe_ip

        cmd = [
            "nmap",
            "-sV",
            "-sC",
            "-O",
            "--system-dns",
            "-Pn",
            "--open",
            "--max-retries", "1",
            "--host-timeout", f"{timeout}s",
            "-p", ports_arg,
            "-oA", str(oA_base),
            ip,
        ]

        result_path = None
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            with _procs_lock:
                _procs.append(proc)
            killed_by_python = False
            try:
                proc.wait(timeout=timeout + 30)
            except subprocess.TimeoutExpired:
                proc.kill()
                proc.wait()
                killed_by_python = True
            finally:
                with _procs_lock:
                    try:
                        _procs.remove(proc)
                    except ValueError:
                        pass
            xml_path = Path(str(oA_base) + ".xml")
            if xml_path.exists():
                if killed_by_python:
                    # nmap was force-killed; the XML file may be truncated/corrupt.
                    # Validate before including — a corrupt file would silently
                    # drop the host from the merged results anyway, but this way
                    # it's counted as a timeout rather than a success.
                    try:
                        ET.parse(xml_path)
                        result_path = xml_path
                    except ET.ParseError:
                        xml_path.unlink(missing_ok=True)
                        result_path = None  # counted as failed/timed-out below
                else:
                    result_path = xml_path
        except Exception as e:
            if debug:
                with print_lock:
                    print(f"\n  [!] {ip} error: {e}")

        if not _abort.is_set():
            if sp is not None:
                sp.update()
            if debug:
                with print_lock:
                    status = f"✓ {len(ports)} ports" if result_path else "✗ timeout"
                    print(f"  {ip:<20} {status}")

        return ip, result_path

    def _kill_all_procs() -> None:
        """Terminate all running nmap child processes immediately."""
        with _procs_lock:
            for proc in _procs:
                try:
                    proc.terminate()
                except Exception:
                    pass
        # give them a moment, then force-kill any stragglers
        time.sleep(0.5)
        with _procs_lock:
            for proc in _procs:
                try:
                    if proc.poll() is None:
                        proc.kill()
                except Exception:
                    pass

    results: list[tuple[str, Path | None]] = []
    try:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(scan_host, ip): ip for ip in sorted_hosts}
            for future in as_completed(futures):
                results.append(future.result())
    except KeyboardInterrupt:
        _abort.set()
        _kill_all_procs()
        if sp is not None:
            sp.stop()
        raise

    successful = [(ip, p) for ip, p in results if p is not None]
    failed     = [ip for ip, p in results if p is None]

    combined_xml = outdir / "nmap_combined.xml"
    _merge_nmap_xmls([p for _, p in successful], combined_xml)

    if sp is not None:
        sp.stop(f"  [+] nmap: {len(successful)}/{total} complete"
              + (f", {len(failed)} timed out" if failed else ""))
    else:
        print(f"  [+] nmap: {len(successful)}/{total} complete"
              + (f", {len(failed)} timed out" if failed else ""))

    # Write timed-out hosts to debug dir for reference
    if failed:
        debug_dir = outdir / "debug"
        debug_dir.mkdir(exist_ok=True)
        with open(debug_dir / "nmap_timed_out.txt", "w") as f:
            for ip in sort_ips(failed):
                f.write(ip + "\n")

    return combined_xml, failed



# ----------------------------------------------------------------------------- #
# Stage 5: XLSX report
# ----------------------------------------------------------------------------- #
def write_xlsx(combined_xml: Path, outdir: Path) -> Path | None:
    """
    Parse nmap combined XML and write nmap_results.xlsx.
    Columns: IP | Hostname | MAC | Vendor | OS | Port | Protocol | State | Service | Product | Version | Extra Info
    One row per port entry in the XML. Timed-out hosts are not included
    as they do not appear in nmap_combined.xml (only successful scans are merged).
    """
    if not HAS_OPENPYXL:
        print("  [!] openpyxl not installed — skipping xlsx")
        print("      install: python3 -m venv .venv && source .venv/bin/activate && pip install openpyxl")
        return None

    try:
        tree = ET.parse(combined_xml)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"  [!] xlsx: failed to parse {combined_xml}: {e}")
        return None

    HEADER = ["IP", "Hostname", "MAC", "Vendor", "OS", "Port", "Protocol",
              "State", "Service", "Product", "Version", "Extra Info"]

    # Header style
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Alternating row fills
    FILL_ODD  = PatternFill("solid", fgColor="DCE6F1")
    FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "nmap results"
    ws.freeze_panes = "A2"

    # Write header
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN

    row_idx = 1  # track for alternating fills

    for host in root.findall("host"):
        # ── addresses ────────────────────────────────────────────────────────
        ip      = "unknown"
        mac     = ""
        vendor  = ""
        for addr in host.findall("address"):
            atype = addr.get("addrtype", "")
            aval  = addr.get("addr", "")
            if atype in ("ipv4", "ipv6") and ip == "unknown":
                ip = aval
            elif atype == "mac":
                mac    = aval
                vendor = addr.get("vendor", "")

        # ── hostname ──────────────────────────────────────────────────────────
        hostname = ""
        for hn in host.findall(".//hostname"):
            name = hn.get("name", "")
            if name:
                hostname = name
                break

        # ── OS ───────────────────────────────────────────────────────────────
        os_match = host.find(".//osmatch")
        os_name  = os_match.get("name", "") if os_match is not None else ""

        # ── ports ─────────────────────────────────────────────────────────────
        for port in host.findall(".//port"):
            state_el   = port.find("state")
            service_el = port.find("service")

            port_id  = port.get("portid", "")
            protocol = port.get("protocol", "tcp")
            state    = state_el.get("state",   "") if state_el   is not None else ""
            svc      = service_el.get("name",    "") if service_el is not None else ""
            product  = service_el.get("product", "") if service_el is not None else ""
            version  = service_el.get("version", "") if service_el is not None else ""
            extra    = service_el.get("extrainfo", "") if service_el is not None else ""

            row_idx += 1
            fill = FILL_ODD if row_idx % 2 else FILL_EVEN
            row  = [ip, hostname, mac, vendor, os_name,
                    port_id, protocol, state, svc, product, version, extra]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # ── column widths — dynamic based on content, capped to avoid absurd widths ─
    MIN_WIDTH = 8
    MAX_WIDTH = 60
    PADDING   = 3
    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        width = min(MAX_WIDTH, max(MIN_WIDTH, max_len + PADDING))
        ws.column_dimensions[col_cells[0].column_letter].width = width

    # ── auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    xlsx_path = outdir / "nmap_results.xlsx"
    wb.save(xlsx_path)
    total_rows = ws.max_row - 1  # exclude header
    print(f"  [+] xlsx: {total_rows} rows → {xlsx_path}")
    return xlsx_path


# ----------------------------------------------------------------------------- #
# Stage 6: EyeWitness
# ----------------------------------------------------------------------------- #
def run_eyewitness(combined_xml: Path, outdir: Path) -> Path | None:
    """
    Run EyeWitness against the merged nmap XML.
    Returns path to report.html if successful, None otherwise.
    """
    ew_dir = outdir / "eyewitness"
    if ew_dir.exists():
        shutil.rmtree(ew_dir)

    cmd = [
        "eyewitness",
        "-x", str(combined_xml),
        "--web",
        "--no-prompt",
        "-d", str(ew_dir),
    ]

    print(f"  [*] running EyeWitness...", end="", flush=True)
    result = subprocess.run(cmd, capture_output=True)
    report = ew_dir / "report.html"

    if result.returncode != 0 or not report.exists():
        print(f"\r  [!] EyeWitness failed (exit {result.returncode})")
        return None

    print(f"\r  [+] EyeWitness done")
    return report


# ----------------------------------------------------------------------------- #
# Main
# ----------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "scanner.py — masscan + nmap + EyeWitness pipeline\n"
            "by x41 and the praktikant\n"
            "\n"
            "Stages:\n"
            "  1. masscan   full TCP port discovery (1-65535)\n"
            "  2. parse     extract hosts/ports from masscan XML\n"
            "  3. confirm   full TCP connect per port, filters proxied/dead ports\n"
            "               only confirmed-open ports proceed to nmap\n"
            "  4. nmap      -sV -sC -O per host on confirmed ports only\n"
            "  5. xlsx      nmap_results.xlsx from combined XML\n"
            "  6. eyewitness screenshots web services (optional, if installed)\n"
            "\n"
            "Defaults: --rate 50000  --workers 20  --timeout 400\n"
            "          --confirm-workers 25  --confirm-timeout 2\n"
            "\n"
            "Examples:\n"
            "  sudo python3 scanner.py 192.168.1.0/24\n"
            "  sudo python3 scanner.py --iL targets.txt --rate 10000\n"
            "  sudo python3 scanner.py 10.0.0.0/8 --exclude 10.0.0.1\n"
            "  sudo python3 scanner.py 10.0.0.0/16 --workers 30\n"
            "  sudo python3 scanner.py --from-masscan-xml masscan.xml --outdir results/\n"
            "  sudo python3 scanner.py --from-nmap-xml nmap_combined.xml\n"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # Target (mutually exclusive: positional or --iL)
    target_group = parser.add_argument_group("target (provide one)")
    target_group.add_argument(
        "target",
        nargs="?",
        default=None,
        metavar="TARGET",
        help="IP address or CIDR  (e.g. 192.168.1.0/24, 10.0.0.0/8)"
    )
    target_group.add_argument(
        "--iL",
        dest="il_file",
        default=None,
        metavar="FILE",
        help="File containing one target per line (IPs, CIDRs)"
    )

    # Scan options
    scan_group = parser.add_argument_group("scan options")
    scan_group.add_argument(
        "--rate",
        type=int,
        default=50_000,
        metavar="PPS",
        help="masscan packet rate in packets/sec  (default: 50000)"
    )
    scan_group.add_argument(
        "--exclude",
        default="",
        metavar="CIDR|FILE",
        help="Exclude a CIDR range or a file of ranges from the scan\n"
             "examples: 192.168.1.1  |  10.0.0.0/8  |  /path/to/exclude.txt"
    )
    scan_group.add_argument(
        "--workers",
        type=int,
        default=20,
        metavar="N",
        help="Parallel nmap deep-dive workers  (default: 20)"
    )
    scan_group.add_argument(
        "--timeout",
        type=int,
        default=400,
        metavar="SEC",
        help="nmap per-host timeout in seconds  (default: 400)"
    )
    scan_group.add_argument(
        "--confirm-workers",
        type=int,
        default=25,
        dest="confirm_workers",
        metavar="N",
        help="Parallel workers for TCP connect confirmation  (default: 25)"
    )
    scan_group.add_argument(
        "--confirm-timeout",
        type=int,
        default=2,
        dest="confirm_timeout",
        metavar="SEC",
        help="Per-host timeout for TCP connect confirmation  (default: 2)"
    )

    # Output options
    out_group = parser.add_argument_group("output options")
    out_group.add_argument(
        "--outdir",
        default=None,
        metavar="DIR",
        help="Output directory  (default: scan_<target>_<timestamp>/)"
    )

    # Developer options
    dev_group = parser.add_argument_group("developer options")
    dev_group.add_argument(
        "--from-masscan-xml",
        default=None,
        metavar="MASSCAN_XML",
        help="Skip masscan stage, use existing masscan XML file\n"
             "useful for re-running parse/confirm/nmap without rescanning"
    )
    dev_group.add_argument(
        "--from-nmap-xml",
        default=None,
        metavar="NMAP_XML",
        help="Skip all scan stages, use existing nmap combined XML file\n"
             "runs only xlsx + eyewitness — useful for re-reporting"
    )
    dev_group.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="Show verbose per-host output during confirm and nmap stages"
    )

    return parser.parse_args()


def main() -> None:
    # Suppress default SIGINT traceback — we handle it cleanly at __main__
    signal.signal(signal.SIGINT, lambda *_: (_ for _ in ()).throw(KeyboardInterrupt()))

    args = parse_args()

    # masscan and nmap require root; report-only modes do not
    if not args.from_nmap_xml and os.geteuid() != 0:
        print("[!] Requires root. Run: sudo python3 scanner.py ...")
        sys.exit(1)

    # --from-nmap-xml: skip everything except xlsx + eyewitness
    if args.from_nmap_xml:
        nmap_xml = Path(args.from_nmap_xml)
        if not nmap_xml.exists():
            print(f"[!] nmap XML not found: {nmap_xml}")
            sys.exit(1)
        outdir = Path(args.outdir) if args.outdir else nmap_xml.parent
        outdir.mkdir(parents=True, exist_ok=True)
        optional             = check_dependencies(skip_masscan=True, require_nmap=False)
        run_eyewitness_stage = optional.get("eyewitness", False)
        print(f"[*] --from-nmap-xml  : skipping scan stages")
        print(f"[*] nmap XML         : {nmap_xml}")
        print(f"[*] Output dir       : {outdir}/")
        print()
        print("[5/6] xlsx     — building results spreadsheet")
        xlsx_path = write_xlsx(nmap_xml, outdir)
        ew_report = None
        if run_eyewitness_stage:
            print("[6/6] eyewitness — web screenshots")
            ew_report = run_eyewitness(nmap_xml, outdir)
        else:
            print("[6/6] eyewitness — skipped (not installed)")
        print()
        print("=" * 60)
        print("  REPORT COMPLETE")
        print("=" * 60)
        if xlsx_path:
            print(f"  xlsx report     : {xlsx_path}")
        if ew_report:
            print(f"  eyewitness      : {ew_report}")
        elif run_eyewitness_stage:
            print(f"  eyewitness      : failed — check {outdir}/eyewitness/")
        else:
            print(f"  eyewitness      : run manually:")
            print(f"                    eyewitness -x {nmap_xml} --web --no-prompt -d {outdir}/eyewitness")
        print("=" * 60)
        print()
        sudo_user = os.environ.get("SUDO_USER")
        if sudo_user:
            subprocess.run(["chown", "-R", f"{sudo_user}:{sudo_user}", str(outdir)],
                           capture_output=True)
        sys.exit(0)

    debug = args.debug

    scan_start = datetime.now()

    if args.from_masscan_xml:
        if not os.path.isfile(args.from_masscan_xml):
            print(f"[!] XML file not found: {args.from_masscan_xml}")
            sys.exit(1)
        optional             = check_dependencies(skip_masscan=True)
        run_eyewitness_stage = optional.get("eyewitness", False)
        masscan_xml    = Path(args.from_masscan_xml)
        outdir         = Path(args.outdir) if args.outdir else masscan_xml.parent
        display_target = str(masscan_xml)
        outdir.mkdir(parents=True, exist_ok=True)
        print(f"[*] --from-masscan-xml mode  : skipping masscan, using {masscan_xml}")
        print(f"[*] Output dir       : {outdir}/")
        print()

    else:
        if not args.il_file and not args.target:
            print("[!] Provide a target or --iL <file>")
            sys.exit(1)
        if args.il_file and not os.path.isfile(args.il_file):
            print(f"[!] File not found: {args.il_file}")
            sys.exit(1)
        if args.target:
            validate_target(args.target)

        # All targets validated — now check dependencies
        optional             = check_dependencies(skip_masscan=False)
        run_eyewitness_stage = optional.get("eyewitness", False)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.outdir:
            outdir = Path(args.outdir)
        else:
            safe   = re.sub(r"[/:\ ]", "_", args.il_file or args.target)
            safe   = re.sub(r"[^A-Za-z0-9._-]", "", safe)
            outdir = Path(f"scan_{safe}_{ts}")

        outdir.mkdir(parents=True, exist_ok=True)
        masscan_xml = outdir / f"masscan_{ts}.xml"

        if args.il_file:
            host_count     = estimate_hosts_from_file(args.il_file)
            display_target = f"iL:{args.il_file}"
        else:
            host_count     = estimate_hosts(args.target)
            display_target = args.target

        est_secs = (65535 * host_count) // args.rate

        print(f"  target          : {display_target}")
        print(f"  rate            : {args.rate:,} pps")
        print(f"  workers         : nmap={args.workers}  confirm={args.confirm_workers}")
        print(f"  timeout         : nmap={args.timeout}s  confirm={args.confirm_timeout}s")
        print(f"  est. time       : ~{format_time(est_secs)}")
        print(f"  output          : {outdir}/")
        if args.exclude:
            print(f"  exclude         : {args.exclude}")
        if debug:
            print(f"  debug           : enabled")
        print()

        print("[1/6] masscan — full port discovery")
        run_masscan(
            rate    = args.rate,
            target  = args.target or "",
            il_file = args.il_file or "",
            exclude = args.exclude,
            outfile = masscan_xml,
        )

        if not masscan_xml.exists() or masscan_xml.stat().st_size == 0:
            print("[!] masscan produced no output.")
            sys.exit(1)

    print("[2/6] parse   — extracting hosts and ports")
    portmap = parse_masscan_xml(masscan_xml)

    if not portmap:
        print("[!] No open TCP ports found.")
        sys.exit(0)

    portmap_path = write_parse_outputs(portmap, outdir)

    print("[3/6] confirm — TCP connect, filtering proxied/dead ports")
    confirmed_map, confirmed_ports_path = confirm_hosts(
        portmap  = portmap,
        workers  = args.confirm_workers,
        timeout  = args.confirm_timeout,
        outdir   = outdir,
        debug    = debug,
    )
    if not confirmed_map:
        print("[!] No hosts survived confirmation.")
        sys.exit(0)

    print("[4/6] nmap    — deep dive fingerprinting")
    combined_xml, timed_out = run_nmap_deep_dive(
        confirmed_map = confirmed_map,
        workers       = args.workers,
        timeout       = args.timeout,
        outdir        = outdir,
        debug         = debug,
    )

    print("[5/6] xlsx     — building results spreadsheet")
    xlsx_path = write_xlsx(combined_xml, outdir)

    ew_report = None
    if run_eyewitness_stage:
        print("[6/6] eyewitness — web screenshots")
        ew_report = run_eyewitness(combined_xml, outdir)
    else:
        print("[6/6] eyewitness — skipped (not installed)")

    elapsed = datetime.now() - scan_start

    sudo_user = os.environ.get("SUDO_USER")
    if sudo_user:
        subprocess.run(["chown", "-R", f"{sudo_user}:{sudo_user}", str(outdir)],
                       capture_output=True)

    # ── Final summary ────────────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  SCAN COMPLETE")
    print("=" * 60)
    print(f"  target          : {display_target}")
    print(f"  elapsed         : {str(elapsed).split('.')[0]}")
    print(f"  hosts alive     : {len(confirmed_map)}")
    if timed_out:
        print(f"  timed out       : {len(timed_out)}  (see debug/nmap_timed_out.txt)")
    print()
    print("  --- important files ---")
    print(f"  nmap XML        : {combined_xml}")
    print(f"  nmap per-host   : {outdir}/nmap/  (.xml .nmap .gnmap)")
    if xlsx_path:
        print(f"  xlsx report     : {xlsx_path}")
    else:
        print(f"  xlsx report     : skipped")
        print(f"                    install: python3 -m venv .venv && source .venv/bin/activate && pip install openpyxl")
    if ew_report:
        print(f"  eyewitness      : {ew_report}")
    elif run_eyewitness_stage:
        print(f"  eyewitness      : failed — check {outdir}/eyewitness/")
    else:
        print(f"  eyewitness      : run manually:")
        print(f"                    eyewitness -x {combined_xml} --web --no-prompt -d {outdir}/eyewitness")
    print()
    print("  --- debug / raw data ---")
    print(f"  masscan ports   : {portmap_path}")
    print(f"  confirmed ports : {confirmed_ports_path}")
    print(f"  {outdir}/debug/  (hosts, summaries, timed-out)")
    print("=" * 60)
    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        # Spinner may be mid-print — erase the line before writing anything
        print(f"\r{chr(32) * 60}\r", end="")
        print()
        print("[!] Aborted.")
        sys.exit(130)  # standard exit code for SIGINT#!/usr/bin/env python3
# NOT SAFE FOR WORK YET!
# ----------------------------------------------------------------------------- #
#                    scanner.py — masscan + parse + confirm + nmap + eyewitness #
#                         by x41 and the praktikant  (v2.1)                     #
# ----------------------------------------------------------------------------- #
# Stages:
#   1. masscan     — full port discovery (1-65535)
#   2. parse       — extract hosts/ports from masscan XML
#   3. confirm     — parallel nmap TCP connect, filters SYN proxy false positives
#   4. deep        — nmap -sV -sC -O per confirmed host on confirmed ports
#                    produces per-host XMLs + merged nmap_combined.xml
#   5. eyewitness  — screenshots web services from combined nmap XML (-x)
#                    correct service labels guaranteed by scaled timeouts
#                    flat timeout per profile ensures nmap finishes cleanly
#
# Usage:
#   sudo python3 scanner.py --profile <safe|normal|fast|lab> <TARGET>
#   sudo python3 scanner.py --profile <safe|normal|fast|lab> --iL <FILE>
#
# Skip masscan, reuse existing XML (for testing/dev):
#   sudo python3 scanner.py --profile normal --from-xml <masscan.xml>
# ----------------------------------------------------------------------------- #

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import threading
import xml.etree.ElementTree as ET
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from ipaddress import ip_address, ip_network
from pathlib import Path
from threading import Lock

# ----------------------------------------------------------------------------- #
# Profiles
# ----------------------------------------------------------------------------- #
PROFILES = {
    "safe": {
        "rate":            10_000,
        "retries":         0,
        "confirm":         True,
        "bw":              "~3.2 Mbit/s",
        "desc":            "healthcare / critical infra",
        "confirm_workers": 10,
        "confirm_timeout": 5,
        "nmap_workers":    3,
        "nmap_timeout":    300,  # flat — iLO/embedded devices slow regardless of port count
    },
    "normal": {
        "rate":            50_000,
        "retries":         1,
        "confirm":         False,
        "bw":              "~16 Mbit/s",
        "desc":            "standard corporate / SMB",
        "confirm_workers": 25,
        "confirm_timeout": 2,
        "nmap_workers":    20,
        "nmap_timeout":    400,  # flat — 2x real-world max observed (182s)
    },
    "fast": {
        "rate":            100_000,
        "retries":         1,
        "confirm":         False,
        "bw":              "~32 Mbit/s",
        "desc":            "large orgs / pentest VLAN",
        "confirm_workers": 50,
        "confirm_timeout": 2,
        "nmap_workers":    30,
        "nmap_timeout":    400,
    },
    "lab": {
        "rate":            500_000,
        "retries":         2,
        "confirm":         False,
        "bw":              "~160 Mbit/s",
        "desc":            "your lab only — never on customer networks",
        "confirm_workers": 100,
        "confirm_timeout": 1,
        "nmap_workers":    30,
        "nmap_timeout":    400,
    },
}

PROBE_PREFER = [22, 445, 80, 443, 389, 3389, 21, 25, 8080, 8443, 8888, 9090, 9443, 5986]


# ----------------------------------------------------------------------------- #
# Progress bar
# ----------------------------------------------------------------------------- #
def _bar(n: int, total: int, width: int = 35, label: str = "") -> str:
    """Return an animated progress bar string for in-place terminal updates."""
    pct   = n / total if total else 1
    filled = int(width * pct)
    bar   = "█" * filled + "░" * (width - filled)
    suffix = f" {label}" if label else ""
    return f"  [{bar}] {n}/{total}{suffix}"


def _print_bar(n: int, total: int, label: str = "") -> None:
    """Print progress bar in-place (overwrites current line)."""
    print(f"\r{_bar(n, total, label=label)}", end="", flush=True)


def _end_bar(total: int, label: str = "") -> None:
    """Complete the bar at 100% and move to next line."""
    print(f"\r{_bar(total, total, label=label)}")


# ----------------------------------------------------------------------------- #
# Helpers
# ----------------------------------------------------------------------------- #
def sort_ips(ips: list[str]) -> list[str]:
    return sorted(ips, key=lambda x: ip_address(x))


def estimate_hosts(target: str) -> int:
    try:
        net = ip_network(target, strict=False)
        n = net.num_addresses
        return max(1, n - 2) if n > 2 else n
    except ValueError:
        return 1


def estimate_hosts_from_file(path: str) -> int:
    total = 0
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            total += estimate_hosts(line)
    return total


def format_time(seconds: int) -> str:
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"


def select_probe(ports: set[int]) -> int:
    for p in PROBE_PREFER:
        if p in ports:
            return p
    return min(ports)


def validate_target(target: str) -> None:
    try:
        net = ip_network(target, strict=False)
        if net.num_addresses == 1:
            print(f"[!] Target {target} is a single host (/32).")
            print(f"    masscan on a single host may miss ports due to self-routing load.")
            print(f"    For single hosts, nmap is more reliable.")
            answer = input("    Continue anyway? (yes/no): ").strip()
            if answer != "yes":
                print("[!] Aborted.")
                sys.exit(1)
    except ValueError:
        print(f"[!] Invalid target: {target}")
        sys.exit(1)


def confirm_safe_profile() -> None:
    print("  ┌─────────────────────────────────────────────────────────┐")
    print("  │  SAFE profile active. Before continuing confirm:        │")
    print("  │  • Written authorization obtained for this target       │")
    print("  │  • Fragile / medical devices identified and excluded    │")
    print("  └─────────────────────────────────────────────────────────┘")
    print()
    answer = input("  Type 'yes' to continue: ").strip()
    if answer != "yes":
        print("[!] Aborted.")
        sys.exit(1)
    print()


# ----------------------------------------------------------------------------- #
# Dependency checks
# ----------------------------------------------------------------------------- #
def check_dependencies(skip_masscan: bool) -> dict[str, bool]:
    required = ["nmap"]
    if not skip_masscan:
        required.append("masscan")

    for tool in required:
        found = subprocess.run(["which", tool], capture_output=True).returncode == 0
        if not found:
            print(f"[!] required tool not found: {tool}")
            sys.exit(1)

    optional = {"eyewitness": False}
    for tool in optional:
        found = subprocess.run(["which", tool], capture_output=True).returncode == 0
        optional[tool] = found
        if not found:
            answer = input(f"[?] {tool} not installed — continue without it? (yes/no): ").strip()
            if answer != "yes":
                print("[!] Aborted.")
                sys.exit(1)

    return optional


# ----------------------------------------------------------------------------- #
# Stage 1: masscan
# ----------------------------------------------------------------------------- #
def run_masscan(
    profile:    dict,
    target:     str,
    il_file:    str,
    port_range: str,
    exclude:    str,
    outfile:    Path,
) -> None:
    cmd = [
        "masscan",
        "-p", port_range,
        "--rate", str(profile["rate"]),
        "--retries", str(profile["retries"]),
        "--open-only",
        "-oX", str(outfile),
    ]

    if il_file:
        cmd += ["-iL", il_file]
    else:
        cmd.append(target)

    if exclude:
        if os.path.isfile(exclude):
            cmd += ["--excludefile", exclude]
        else:
            cmd += ["--exclude", exclude]

    # Run masscan with stderr piped so we can parse its progress lines.
    # masscan writes status to stderr in the form:
    #   rate:  49983; found: 234; remaining: 00:01:22; waiting...
    # We parse "remaining" to show a live progress bar.
    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        text=True,
    )

    found_count = {"n": 0}
    rate_str    = {"v": ""}
    done_event  = threading.Event()

    def read_stderr():
        for line in proc.stderr:
            line = line.strip()
            # rate:  49983; found: 234; remaining: 00:01:22; waiting...
            m_found = re.search(r"found:\s*(\d+)", line)
            m_rate  = re.search(r"rate:\s*([\d,]+)", line)
            m_rem   = re.search(r"remaining:\s*(\S+)", line)
            if m_found:
                found_count["n"] = int(m_found.group(1))
            if m_rate:
                rate_str["v"] = m_rate.group(1).replace(",", "")
            if m_rem:
                rem = m_rem.group(1)
                label = f"found: {found_count['n']}  remaining: {rem}  rate: {rate_str['v']} pps"
                print(f"\r  [masscan] {label:<70}", end="", flush=True)
        done_event.set()

    t = threading.Thread(target=read_stderr, daemon=True)
    t.start()
    proc.wait()
    done_event.wait(timeout=5)
    print(f"\r  [+] masscan done — {found_count['n']} open ports found{' ' * 40}")

    if proc.returncode not in (0, 1):  # masscan exits 1 on Ctrl+C partial runs
        print(f"[!] masscan exited with code {proc.returncode}")
        sys.exit(1)


# ----------------------------------------------------------------------------- #
# Stage 2: parse masscan XML
# ----------------------------------------------------------------------------- #
def parse_masscan_xml(path: Path) -> dict[str, set[int]]:
    try:
        tree = ET.parse(path)
    except ET.ParseError as e:
        print(f"[!] Failed to parse masscan XML: {e}")
        sys.exit(1)

    root = tree.getroot()
    portmap: dict[str, set[int]] = defaultdict(set)

    for host in root.findall("host"):
        addr_el = host.find("address[@addrtype='ipv4']")
        if addr_el is None:
            continue
        ip = addr_el.get("addr")
        if not ip:
            continue
        for port_el in host.findall(".//port[@protocol='tcp']"):
            state_el = port_el.find("state[@state='open']")
            if state_el is None:
                continue
            portid = port_el.get("portid")
            if portid and portid.isdigit():
                portmap[ip].add(int(portid))

    return portmap


def write_parse_outputs(portmap: dict[str, set[int]], outdir: Path) -> Path:
    """Write masscan parse outputs. Returns portmap_path for summary use."""
    sorted_hosts = sort_ips(list(portmap.keys()))
    all_ports    = sorted(set(p for ports in portmap.values() for p in ports))
    probes       = {ip: select_probe(portmap[ip]) for ip in sorted_hosts}

    # Important output — confirmed ports map used by nmap stage
    portmap_path = outdir / "masscan_ports_map.csv"
    with open(portmap_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["host", "ports"])
        for ip in sorted_hosts:
            writer.writerow([ip, ",".join(str(p) for p in sorted(portmap[ip]))])

    # Debug outputs — useful for troubleshooting, not needed day-to-day
    debug_dir = outdir / "debug"
    debug_dir.mkdir(exist_ok=True)

    with open(debug_dir / "masscan_alive_hosts.txt", "w") as f:
        for ip in sorted_hosts:
            f.write(ip + "\n")

    with open(debug_dir / "masscan_union_ports.txt", "w") as f:
        for port in all_ports:
            f.write(str(port) + "\n")

    # Note: probe column is the "best" port per host for manual re-probing reference.
    # The confirmation stage itself runs against all masscan ports, not just this one.
    with open(debug_dir / "masscan_host_best_ports.txt", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["host", "best_port"])
        for ip in sorted_hosts:
            writer.writerow([ip, probes[ip]])

    lines = [
        "=== masscan parse summary ===",
        "",
        f"Alive hosts (masscan)  : {len(sorted_hosts)}",
        f"Unique ports           : {len(all_ports)}",
        "",
        "--- Per-host breakdown (ports / probe) ---",
    ]
    for ip in sorted_hosts:
        ports_str = ",".join(str(p) for p in sorted(portmap[ip]))
        lines.append(f"  {ip:<20}  ports: {ports_str:<60}  probe: {probes[ip]}")

    with open(debug_dir / "masscan_summary.txt", "w") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  [+] masscan: {len(sorted_hosts)} hosts, {len(all_ports)} unique ports")
    return portmap_path


# ----------------------------------------------------------------------------- #
# Stage 3: parallel nmap TCP connect confirmation
# ----------------------------------------------------------------------------- #
def parse_nmap_open_ports(xml_output: str) -> set[int]:
    confirmed = set()
    try:
        root = ET.fromstring(xml_output)
        for port_el in root.findall(".//port"):
            state_el = port_el.find("state[@state='open']")
            if state_el is not None:
                portid = port_el.get("portid")
                if portid and portid.isdigit():
                    confirmed.add(int(portid))
    except ET.ParseError:
        pass
    return confirmed


def confirm_single_host(
    ip:      str,
    ports:   set[int],
    timeout: int,
) -> tuple[str, set[int]]:
    ports_arg = ",".join(str(p) for p in sorted(ports))
    cmd = [
        "nmap",
        "-sT", "-Pn", "-n",
        "--open",
        "--max-retries", "0",
        "--host-timeout", f"{timeout}s",
        "-p", ports_arg,
        "-oX", "-",
        ip,
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout + 5,
        )
        return ip, parse_nmap_open_ports(result.stdout)
    except subprocess.TimeoutExpired:
        return ip, set()
    except Exception:
        return ip, set()


def confirm_hosts(
    portmap: dict[str, set[int]],
    profile: dict,
    outdir:  Path,
    debug:   bool = False,
) -> dict[str, set[int]]:
    sorted_hosts   = sort_ips(list(portmap.keys()))
    total          = len(sorted_hosts)
    workers        = profile["confirm_workers"]
    timeout        = profile["confirm_timeout"]
    confirmed_map: dict[str, set[int]] = {}
    dropped:       list[str] = []

    counter      = {"n": 0}
    counter_lock = Lock()
    print_lock   = Lock()

    _print_bar(0, total)

    def task(ip: str) -> tuple[str, set[int]]:
        return confirm_single_host(ip, portmap[ip], timeout)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(task, ip): ip for ip in sorted_hosts}
        for future in as_completed(futures):
            ip, confirmed_ports = future.result()
            with counter_lock:
                counter["n"] += 1
                n = counter["n"]
            with print_lock:
                if confirmed_ports:
                    confirmed_map[ip] = confirmed_ports
                    if debug:
                        conf_str = ",".join(str(p) for p in sorted(confirmed_ports))
                        print(f"\n  {ip:<20} ✓  {conf_str}")
                else:
                    dropped.append(ip)
                    if debug:
                        print(f"\n  {ip:<20} ✗  dropped")
                _print_bar(n, total)

    _end_bar(total)

    confirmed_sorted    = sort_ips(list(confirmed_map.keys()))
    all_confirmed_ports = sorted(set(p for ports in confirmed_map.values() for p in ports))

    # Always write confirmed ports map — important output used by nmap stage
    confirmed_ports_path = outdir / "confirmed_ports_map.csv"
    with open(confirmed_ports_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["host", "ports"])
        for ip in confirmed_sorted:
            writer.writerow([ip, ",".join(str(p) for p in sorted(confirmed_map[ip]))])

    # Debug outputs
    debug_dir = outdir / "debug"
    debug_dir.mkdir(exist_ok=True)

    with open(debug_dir / "confirmed_hosts.txt", "w") as f:
        for ip in confirmed_sorted:
            f.write(ip + "\n")

    lines = [
        "=== confirmation summary ===",
        "",
        f"Hosts in (masscan)     : {total}",
        f"Hosts confirmed alive  : {len(confirmed_sorted)}",
        f"Hosts dropped          : {len(dropped)}",
        f"Unique confirmed ports : {len(all_confirmed_ports)}",
        "",
        "--- Confirmed per-host ---",
    ]
    for ip in confirmed_sorted:
        ports_str = ",".join(str(p) for p in sorted(confirmed_map[ip]))
        lines.append(f"  {ip:<20}  {ports_str}")
    if dropped:
        lines += ["", "--- Dropped (SYN proxy / not alive) ---"]
        for ip in sort_ips(dropped):
            lines.append(f"  {ip}")
    with open(debug_dir / "confirmed_summary.txt", "w") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  [+] confirmed: {len(confirmed_sorted)} alive, {len(dropped)} dropped")
    return confirmed_map


# ----------------------------------------------------------------------------- #
# Stage 4: nmap deep dive
# ----------------------------------------------------------------------------- #
def _merge_nmap_xmls(xml_paths: list[Path], out_path: Path) -> None:
    if not xml_paths:
        return

    host_blocks: list[str] = []
    base_attribs: dict     = {}

    for path in xml_paths:
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            if not base_attribs:
                base_attribs = root.attrib.copy()
            for host_el in root.findall("host"):
                host_blocks.append(ET.tostring(host_el, encoding="unicode"))
        except ET.ParseError:
            continue

    with open(out_path, "w") as f:
        f.write('<?xml version="1.0"?>\n')
        f.write(
            f'<nmaprun scanner="nmap" args="scanner.py combined" '
            f'start="{base_attribs.get("start", "")}" '
            f'version="{base_attribs.get("version", "")}" '
            f'xmloutputversion="{base_attribs.get("xmloutputversion", "1.04")}">\n'
        )
        for block in host_blocks:
            f.write(block + "\n")
        f.write("</nmaprun>\n")


def run_nmap_deep_dive(
    confirmed_map: dict[str, set[int]],
    profile:       dict,
    outdir:        Path,
    debug:         bool = False,
) -> tuple[Path, list[str]]:
    sorted_hosts = sort_ips(list(confirmed_map.keys()))
    total        = len(sorted_hosts)
    workers      = profile["nmap_workers"]
    timeout      = profile["nmap_timeout"]

    nmap_dir = outdir / "nmap"
    nmap_dir.mkdir(exist_ok=True)

    _print_bar(0, total)

    counter      = {"n": 0}
    counter_lock = Lock()
    print_lock   = Lock()

    def scan_host(ip: str) -> tuple[str, Path | None]:
        ports     = confirmed_map[ip]
        ports_arg = ",".join(str(p) for p in sorted(ports))
        safe_ip   = ip.replace(".", "_")
        # -oA produces .xml + .nmap + .gnmap
        oA_base   = nmap_dir / safe_ip

        cmd = [
            "nmap",
            "-sV",
            "-sC",
            "-O",
            "--system-dns",
            "-Pn",
            "--open",
            "--max-retries", "1",
            "--host-timeout", f"{timeout}s",
            "-p", ports_arg,
            "-oA", str(oA_base),
            ip,
        ]

        with counter_lock:
            counter["n"] += 1
            n = counter["n"]

        try:
            subprocess.run(cmd, capture_output=True, timeout=timeout + 30)
            xml_path = Path(str(oA_base) + ".xml")
            result_path = xml_path if xml_path.exists() else None
        except subprocess.TimeoutExpired:
            result_path = None
            if debug:
                with print_lock:
                    print(f"\n  [!] {ip} timed out")
        except Exception as e:
            result_path = None
            if debug:
                with print_lock:
                    print(f"\n  [!] {ip} error: {e}")

        with print_lock:
            _print_bar(n, total)
            if debug:
                status = f"✓ {len(ports)} ports" if result_path else "✗ timeout"
                print(f"\n  {ip:<20} {status}")

        return ip, result_path

    results: list[tuple[str, Path | None]] = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(scan_host, ip): ip for ip in sorted_hosts}
        for future in as_completed(futures):
            results.append(future.result())

    successful = [(ip, p) for ip, p in results if p is not None]
    failed     = [ip for ip, p in results if p is None]

    combined_xml = outdir / "nmap_combined.xml"
    _merge_nmap_xmls([p for _, p in successful], combined_xml)

    _end_bar(total)
    print(f"  [+] nmap: {len(successful)}/{total} complete"
          + (f", {len(failed)} timed out" if failed else ""))

    # Write timed-out hosts to debug dir for reference
    if failed:
        debug_dir = outdir / "debug"
        debug_dir.mkdir(exist_ok=True)
        with open(debug_dir / "nmap_timed_out.txt", "w") as f:
            for ip in sort_ips(failed):
                f.write(ip + "\n")

    return combined_xml, failed


# ----------------------------------------------------------------------------- #
# Stage 5: EyeWitness
# ----------------------------------------------------------------------------- #
def run_eyewitness(combined_xml: Path, outdir: Path) -> Path | None:
    """
    Run EyeWitness against the merged nmap XML.
    Returns path to report.html if successful, None otherwise.
    """
    ew_dir = outdir / "eyewitness"
    if ew_dir.exists():
        shutil.rmtree(ew_dir)

    cmd = [
        "eyewitness",
        "-x", str(combined_xml),
        "--web",
        "--no-prompt",
        "-d", str(ew_dir),
    ]

    print(f"  [*] running EyeWitness...", end="", flush=True)
    result = subprocess.run(cmd, capture_output=True)
    report = ew_dir / "report.html"

    if result.returncode != 0 or not report.exists():
        print(f"\r  [!] EyeWitness failed (exit {result.returncode})")
        return None

    print(f"\r  [+] EyeWitness done")
    return report


# ----------------------------------------------------------------------------- #
# Main
# ----------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "scanner.py — masscan + nmap + EyeWitness pipeline\n"
            "by x41 and the praktikant\n"
            "\n"
            "Stages:\n"
            "  1. masscan   full TCP port discovery (1-65535)\n"
            "  2. parse     extract hosts/ports from masscan XML\n"
            "  3. confirm   parallel nmap connect, filters SYN proxy false positives\n"
            "  4. nmap      -sV -sC -O per host on confirmed ports only\n"
            "               timeout scales automatically with port count\n"
            "  5. eyewitness screenshots web services (optional, if installed)\n"
            "\n"
            "Profiles:\n"
            "  safe   —  10k pps  ~3.2 Mbit/s  healthcare / critical infra\n"
            "             confirmation stage enabled, lowest worker counts\n"
            "  normal —  50k pps  ~16  Mbit/s  standard corporate / SMB\n"
            "  fast   — 100k pps  ~32  Mbit/s  large orgs / pentest VLAN\n"
            "  lab    — 500k pps  ~160 Mbit/s  your lab only — never on customers\n"
            "\n"
            "Examples:\n"
            "  sudo python3 scanner.py --profile normal 192.168.1.0/24\n"
            "  sudo python3 scanner.py --profile fast --iL targets.txt\n"
            "  sudo python3 scanner.py --profile normal 10.0.0.0/8 --exclude 10.0.0.1\n"
            "  sudo python3 scanner.py --profile normal 10.0.0.0/16 --workers 30\n"
            "  sudo python3 scanner.py --profile safe --iL targets.txt --ports 80,443,8080-8090\n"
            "  sudo python3 scanner.py --profile normal --from-xml masscan.xml --outdir results/\n"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )

    # Target (mutually exclusive: positional or --iL)
    target_group = parser.add_argument_group("target (provide one)")
    target_group.add_argument(
        "target",
        nargs="?",
        default=None,
        metavar="TARGET",
        help="IP address, range or CIDR  (e.g. 192.168.1.0/24, 10.0.0.1-254)"
    )
    target_group.add_argument(
        "--iL",
        dest="il_file",
        default=None,
        metavar="FILE",
        help="File containing one target per line (IPs, CIDRs, ranges)"
    )

    # Scan options
    scan_group = parser.add_argument_group("scan options")
    scan_group.add_argument(
        "--profile",
        required=True,
        choices=PROFILES.keys(),
        metavar="PROFILE",
        help="Scan profile: safe | normal | fast | lab  (see above for details)"
    )
    scan_group.add_argument(
        "--ports",
        default="1-65535",
        metavar="PORTS",
        help="masscan port range  (default: 1-65535)\n"
             "examples: 80,443  |  1-1024  |  1-65535"
    )
    scan_group.add_argument(
        "--exclude",
        default="",
        metavar="CIDR|FILE",
        help="Exclude a CIDR range or a file of ranges from the scan\n"
             "examples: 192.168.1.1  |  10.0.0.0/8  |  /path/to/exclude.txt"
    )
    scan_group.add_argument(
        "--workers",
        type=int,
        default=None,
        metavar="N",
        help="Override nmap deep-dive worker count from profile\n"
             "profile defaults: safe=3  normal=20  fast=30  lab=30\n"
             "higher = faster but more load on target network"
    )

    # Output options
    out_group = parser.add_argument_group("output options")
    out_group.add_argument(
        "--outdir",
        default=None,
        metavar="DIR",
        help="Output directory  (default: scan_<target>_<timestamp>/)"
    )

    # Developer options
    dev_group = parser.add_argument_group("developer options")
    dev_group.add_argument(
        "--from-xml",
        default=None,
        metavar="MASSCAN_XML",
        help="Skip masscan stage, use existing masscan XML file\n"
             "useful for re-running parse/confirm/nmap without rescanning"
    )
    dev_group.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="Show verbose per-host output during confirm and nmap stages"
    )

    return parser.parse_args()


def main() -> None:
    if os.geteuid() != 0:
        print("[!] Requires root. Run: sudo python3 scanner.py ...")
        sys.exit(1)

    args    = parse_args()
    profile = PROFILES[args.profile].copy()

    # --workers overrides the profile's nmap_workers at runtime
    if args.workers is not None:
        if args.workers < 1:
            print("[!] --workers must be >= 1")
            sys.exit(1)
        profile["nmap_workers"] = args.workers

    debug = args.debug

    optional             = check_dependencies(skip_masscan=args.from_xml is not None)
    run_eyewitness_stage = optional.get("eyewitness", False)

    scan_start = datetime.now()

    if args.from_xml:
        if not os.path.isfile(args.from_xml):
            print(f"[!] XML file not found: {args.from_xml}")
            sys.exit(1)
        masscan_xml    = Path(args.from_xml)
        outdir         = Path(args.outdir) if args.outdir else masscan_xml.parent
        display_target = str(masscan_xml)
        outdir.mkdir(parents=True, exist_ok=True)
        print(f"[*] --from-xml mode  : skipping masscan, using {masscan_xml}")
        print(f"[*] Output dir       : {outdir}/")
        print()

    else:
        if not args.il_file and not args.target:
            print("[!] Provide a target or --iL <file>")
            sys.exit(1)
        if args.il_file and not os.path.isfile(args.il_file):
            print(f"[!] File not found: {args.il_file}")
            sys.exit(1)
        if args.target:
            validate_target(args.target)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.outdir:
            outdir = Path(args.outdir)
        else:
            safe   = re.sub(r"[/:\ ]", "_", args.il_file or args.target)
            safe   = re.sub(r"[^A-Za-z0-9._-]", "", safe)
            outdir = Path(f"scan_{safe}_{ts}")

        outdir.mkdir(parents=True, exist_ok=True)
        masscan_xml = outdir / f"masscan_{ts}.xml"

        if args.il_file:
            host_count     = estimate_hosts_from_file(args.il_file)
            display_target = f"iL:{args.il_file}"
        else:
            host_count     = estimate_hosts(args.target)
            display_target = args.target

        est_secs = (65535 * host_count) // profile["rate"]

        print(f"  target   : {display_target}")
        print(f"  profile  : {args.profile} — {profile['desc']}")
        print(f"  rate     : {profile['rate']:,} pps  ({profile['bw']})")
        print(f"  ports    : {args.ports}")
        print(f"  est. time: ~{format_time(est_secs)}")
        print(f"  output   : {outdir}/")
        if args.exclude:
            print(f"  exclude  : {args.exclude}")
        if debug:
            print(f"  debug    : enabled")
        print()

        if profile["confirm"]:
            confirm_safe_profile()

        print("[1/5] masscan — full port discovery")
        run_masscan(
            profile    = profile,
            target     = args.target or "",
            il_file    = args.il_file or "",
            port_range = args.ports,
            exclude    = args.exclude,
            outfile    = masscan_xml,
        )

        if not masscan_xml.exists() or masscan_xml.stat().st_size == 0:
            print("[!] masscan produced no output.")
            sys.exit(1)

    print("[2/5] parse   — extracting hosts and ports")
    portmap = parse_masscan_xml(masscan_xml)

    if not portmap:
        print("[!] No open TCP ports found.")
        sys.exit(0)

    portmap_path = write_parse_outputs(portmap, outdir)

    print("[3/5] confirm — filtering false positives")
    confirmed_map = confirm_hosts(portmap, profile, outdir, debug=debug)

    if not confirmed_map:
        print("[!] No hosts survived confirmation.")
        sys.exit(0)

    print("[4/5] nmap    — deep dive fingerprinting")
    combined_xml, timed_out = run_nmap_deep_dive(confirmed_map, profile, outdir, debug=debug)

    ew_report = None
    if run_eyewitness_stage:
        print("[5/5] eyewitness — web screenshots")
        ew_report = run_eyewitness(combined_xml, outdir)
    else:
        print("[5/5] eyewitness — skipped (not installed)")

    elapsed = datetime.now() - scan_start

    sudo_user = os.environ.get("SUDO_USER")
    if sudo_user:
        subprocess.run(["chown", "-R", f"{sudo_user}:{sudo_user}", str(outdir)],
                       capture_output=True)

    # ── Final summary ────────────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  SCAN COMPLETE")
    print("=" * 60)
    print(f"  target          : {display_target}")
    print(f"  elapsed         : {str(elapsed).split('.')[0]}")
    print(f"  hosts alive     : {len(confirmed_map)}")
    if timed_out:
        print(f"  timed out       : {len(timed_out)}  (see debug/nmap_timed_out.txt)")
    print()
    print("  --- important files ---")
    print(f"  confirmed ports : {portmap_path}")
    print(f"  nmap XML        : {combined_xml}")
    print(f"  nmap per-host   : {outdir}/nmap/  (.xml .nmap .gnmap)")
    if ew_report:
        print(f"  eyewitness      : {ew_report}")
    elif run_eyewitness_stage:
        print(f"  eyewitness      : failed — check {outdir}/eyewitness/")
    else:
        print(f"  eyewitness      : run manually:")
        print(f"                    eyewitness -x {combined_xml} --web --no-prompt -d {outdir}/eyewitness")
    print()
    print("  --- debug / raw data ---")
    print(f"  {outdir}/debug/")
    print("=" * 60)
    print()


if __name__ == "__main__":
    main()
